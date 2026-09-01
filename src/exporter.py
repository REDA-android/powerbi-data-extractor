"""
Power BI Data Exporter Module.
Converts extracted datasets into professionally formatted Excel (.xlsx) and CSV (.csv) files.
"""

import io
import pandas as pd
from typing import List, Dict, Any
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class PowerBIExporter:
    """Exports extracted Power BI records to Excel and CSV."""

    @staticmethod
    def to_csv_bytes(records: List[Dict[str, Any]]) -> bytes:
        if not records:
            df = pd.DataFrame([{"Message": "Aucune donnée extraite"}])
        else:
            df = pd.DataFrame(records)
        return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

    @staticmethod
    def to_excel_bytes(records: List[Dict[str, Any]], sheet_name: str = "PowerBI Data") -> bytes:
        if not records:
            df = pd.DataFrame([{"Message": "Aucune donnée extraite"}])
        else:
            df = pd.DataFrame(records)

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Power BI Yellow/Gold & Navy Styling
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="118DFF", end_color="118DFF", fill_type="solid") # Power BI Blue
            thin_border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )

            # Style Header Row
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Style Data Rows and auto-fit columns
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.font = Font(name="Segoe UI", size=10)
                        if cell.row % 2 == 0:
                            cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                    
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)

                worksheet.column_dimensions[col_letter].width = min(max(max_len + 5, 14), 65)

        output.seek(0)
        return output.getvalue()
